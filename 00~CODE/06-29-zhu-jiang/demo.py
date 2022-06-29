import openpyxl
from datetime import datetime
from dateutil.parser import parse

# 计算两个日期的天数差
def get_days(start_date, end_date):
    d1 = parse(start_date)
    d2 = parse(end_date)
    return (d2 - d1).days


# 加载Excel文件
wb = openpyxl.load_workbook(r"d:/temp/随访+门诊对照感觉剖析量表评分统计2.xlsx")

sheet2 = wb.worksheets[1]  # 获取第二个工作表
sheet3 = wb.worksheets[2]  # 获取第三个工作表

# 处理sheet2
for i in range(2, sheet2.max_row + 1):
    value1 = sheet2.cell(i, 5).value
    value2 = sheet2.cell(i, 2).value
    if value1 is None or value2 is None:
        continue
    value1 = str(value1).replace("/", "-").replace(" 00:00:00", "")
    value2 = str(value2).replace("/", "-").replace(" 00:00:00", "")
    days = get_days(value1, value2)
    age = round(days / 365, 3)  # 年龄，保留三位小数
    print(f"出生日期:{value1}, 测试时间:{value2}, 年龄:{age}")
    # 将年龄写入右侧一列
    cell = sheet2.cell(i, 6)
    cell.value = age

# 处理sheet3
for i in range(2, sheet3.max_row + 1):
    value1 = sheet3.cell(i, 5).value
    value2 = sheet3.cell(i, 4).value
    if value1 is None or value2 is None:
        continue
    value1 = str(value1).replace("/", "-").replace(" 00:00:00", "")
    value2 = str(value2).replace("/", "-").replace(" 00:00:00", "")
    days = get_days(value1, value2)
    age = round(days / 365, 3)  # 年龄，保留三位小数
    print(f"出生日期:{value1}, 测试时间:{value2}, 年龄:{age}")
    # 将年龄写入右侧一列
    cell = sheet3.cell(i, 6)
    cell.value = age


# 将结果写入新文件
wb.save(r"d:/temp/result.xlsx")
