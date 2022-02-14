import json
import openpyxl

meals = {}

for filename in ['SZ0001.json', 'SZ0002.json', 'SZ0003.json']:
    f = open(filename, 'r', encoding='gbk')
    meal_class_response = json.load(f)
    meal_classes = meal_class_response['data']['list']
    f.close()
    # 构造菜品字典
    for meal_class in meal_classes:
        print(meal_class['goodsNum'] + "," + meal_class['goodsName'])
        for good in meal_class['goodsCategoryList']:
            meals[good['goodsNum']] = good
            if meal_class['goodsNum'] != good['goodsCategoryNum']:
                print(good)


# 读取数据库导出的数据 T_TAKEAWAY_PLAN_BASE
workbook = openpyxl.load_workbook('take-20210520.xlsx')
sheet = workbook.active
print(sheet.cell(row=1, column=1).value)

with open('update.sql', 'w', encoding='utf8') as f:
    for row_index in range(2, sheet.max_row+1):
        meal_id = sheet.cell(row=row_index, column=6).value
        meal_class_code = sheet.cell(row=row_index, column=7).value
        if meal_id in meals and meal_class_code != meals[meal_id]['goodsCategoryNum']:
            sql = f"update T_TAKEAWAY_PLAN_BASE " \
                  f"set MEAL_CLASS_CODE=\'{meals[meal_id]['goodsCategoryNum']}\', " \
                  f"MEAL_CLASS_NAME=\'{meals[meal_id]['goodsCategoryName']}\' " \
                  f"where ID=\'{meal_id}\';\n"
            f.write(sql)

