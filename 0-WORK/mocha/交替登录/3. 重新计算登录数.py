from openpyxl import load_workbook, Workbook


L = ['10.32.119', '10.32.121', '10.32.124', '10.32.106', '10.32.118', '10.32.111', '10.32.110']
# 新增的ip列表
ip_list = [f'{s}.{i}' for s in L for i in range(1, 254 + 1)]

statistics_path = 'C:/Users/hp/Desktop/交替登陆/需要删除新vpn/result/合并统计信息.xlsx'
detail_path = 'C:/Users/hp/Desktop/交替登陆/需要删除新vpn/result/合并详细信息.xlsx'


statistics_book = load_workbook(statistics_path)
detail_book = load_workbook(detail_path)

statistics_sheet = statistics_book.active
detail_sheet = detail_book.active

statistics_rows = list(statistics_sheet.rows)
detail_rows = list(detail_sheet.rows)


statistics_map = {}
"""
statistics_map
{
    account: {
        total_detail: xxx, # 登录数
        replace_num: xxx # 交替数
    }
}
"""



for index, row in enumerate(detail_rows):
    account, *_, ip = map(lambda c: c.value, row)
    print(f'{account} 登录一次', end=', ')
    if account in statistics_map:
        statistics_map[account]['total_detail'] += 1
    else:
        statistics_map[account] = {
            'total_detail': 0,
            'replace_num': 0
        }
        statistics_map[account]['total_detail'] += 1

    if ip in ip_list:
        print(f'交替登录')
        statistics_map[account]['replace_num'] += 1
    else:
        print()

# 根据统计信息修改表格数据
for index, row in enumerate(statistics_rows):
    # 初始化数据
    if index < 2: continue
    account = row[0].value
    if account in statistics_map:
        # 总登录数 = 原总登录数 - 满足ip的交替登录数
        row[7].value = row[7].value - statistics_map[account]['replace_num']
        # 交替登录数 = 详细列表里对应用户的总数 - 满足ip的登录数
        row[8].value = statistics_map[account]['total_detail'] - statistics_map[account]['replace_num']
    else:
        print(f'找不到 {account} 的数据')
        row[7].value = 0
        row[8].value = 0


# 过滤交替登录数为0的数据
statistics_rows = filter(lambda row: row[8].value != 0, statistics_rows)
# 过滤详细信息中满足ip的数据
detail_rows = filter(lambda row: row[4].value not in ip_list, detail_rows)

wb = Workbook()
ws = wb.active
ws.title = '统计信息'
for index, row in enumerate(statistics_rows):
    for col_index in range(0, 9 + 1):
        ws.cell(row=index + 1, column=col_index + 1, value=row[col_index].value)

ws = wb.create_sheet('详细信息')
for index, row in enumerate(detail_rows):
    for col_index in range(0, 4 + 1):
        ws.cell(row=index + 1, column=col_index + 1, value=row[col_index].value)

wb.save('C:/Users/hp/Desktop/交替登陆/需要删除新vpn/result/处理结果.xlsx')
