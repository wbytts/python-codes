# pip install openpyxl
from openpyxl import load_workbook, Workbook

L = ['10.32.119', '10.32.121', '10.32.124', '10.32.106', '10.32.118', '10.32.111', '10.32.110']
# 新增的ip列表
ip_list = [f'{s}.{i}' for s in L for i in range(1, 254 + 1)]

# 数据源路径
source_dir = 'C:/Users/hp/Desktop/交替登陆/需要删除新vpn/'
# 处理结果存放路径
result_dir = source_dir + 'result/'

final_statistics_rows = []  # 最终的统计信息

# 遍历Excel文件，依次处理
for i in range(1, 24 + 1):
    excel_path = f'{source_dir}{i}交替登录.xlsx'
    print('#####开始处理文件:', excel_path)
    wb = load_workbook(excel_path)
    # 统计信息、详细信息
    statistics_sheet, detail_sheet = wb

    statistics_rows = list(statistics_sheet.rows)  # 统计信息的行数据列表
    detail_rows = list(detail_sheet.rows)  # 详细信息的行数据列表

    need_delete_detail_rows = []  # 存放详细信息中需要删除的行
    result_detail_rows = []  # 存放详细信息删除后的结果
    result_statistics_rows = []  # 存放统计信息修改后的结果

    print('#####开始分析详细信息...')
    for index, row in enumerate(detail_rows):
        if index >= 2:
            account, login_time, replace_account, replace_time, ip = map(lambda c: c.value, row)
            if ip in ip_list:
                need_delete_detail_rows.append(row)
                print(f'{index}\t--->\tip={ip}，在新增列表中')
                continue
        else:
            print(f'{index}\t--->\t非数据行，跳过')
        # 没有匹配到的数据保留下来
        result_detail_rows.append(row)

    need_delete_statistics_indexs = []
    print('#####遍历需要删除的详细信息，更改统计信息...')
    for detail_row in need_delete_detail_rows:
        for index, row in enumerate(statistics_rows):
            if index >= 2 and detail_row[0].value == row[0].value:
                print(f'{row[0].value} 交替登录数减一')
                row[8].value -= 1
                if row[8].value == 0:
                    print(f'{row[0].value} 交替登录数为0，加入待删除列表')
                    need_delete_statistics_indexs.append(index)

    print('#####删除统计信息中交替登录数为0的数据...')
    for index, row in enumerate(statistics_rows):
        if index not in need_delete_statistics_indexs:
            result_statistics_rows.append(row)

    # print('#####更新交替占比...')

    # 存入最终统计信息
    final_statistics_rows.extend(result_statistics_rows)

    # 写入结果信息
    print('#####结果保存中...')
    result_wb = Workbook()
    ws_statistics = result_wb.active
    ws_statistics.title = '统计信息'
    ws_detail = result_wb.create_sheet('详细信息')

    for index, row in enumerate(result_statistics_rows):
        if index == 0:
            ws_statistics.cell(row=index + 1, column=1, value=row[0].value)
            continue
        for col_index in range(0, 9 + 1):
            ws_statistics.cell(row=index + 1, column=col_index + 1, value=row[col_index].value)

    for index, row in enumerate(result_detail_rows):
        if index == 0:
            ws_detail.cell(row=index + 1, column=1, value=row[0].value)
            continue
        for col_index in range(0, 4 + 1):
            ws_detail.cell(row=index + 1, column=col_index + 1, value=row[col_index].value)

    result_wb.save(f'{result_dir}{i}交替登录.xlsx')

# 处理统计信息，合并相同账号的登录数量
row_map = {}
for row in final_statistics_rows:
    account = row[0].value
    login_num = row[7].value
    replace_num = row[8].value
    if account in row_map and login_num is not None and replace_num is not None and isinstance(login_num, int) and isinstance(replace_num, int):
        row_map[account][7].value = row_map[account][7].value + login_num
        row_map[account][8].value = row_map[account][8].value + replace_num
    else:
        row_map[account] = row

# 保存合并结果
wb = Workbook()
ws = wb.active
ws.title = '统计信息'
for index, row in enumerate(row_map.values()):
    for col_index in range(0, 9 + 1):
        ws.cell(row=index + 1, column=col_index + 1, value=row[col_index].value)

wb.save(f'{result_dir}合并统计信息.xlsx')
