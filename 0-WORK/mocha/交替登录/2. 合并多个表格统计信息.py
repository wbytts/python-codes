from openpyxl import load_workbook, Workbook

# 数据源路径
source_dir = 'C:/Users/hp/Desktop/交替登陆/需要删除新vpn'
# 处理结果存放路径
result_dir = source_dir + '/result'


statistics_map = {}

for i in range(1, 24 + 1):
    excel_path = f'{source_dir}/{i}交替登录.xlsx'
    wb = load_workbook(excel_path)
    statistics_sheet = wb['统计信息']
    statistics_rows = list(statistics_sheet.rows)
    for index, row in enumerate(statistics_rows):
        if index > 1:
            values = map(lambda c: c.value, row)
            account, *_, login_num, replace_num, rate = values
            if account == 'liuqinsgs1':
                print(f'{account} --- {login_num} --- {replace_num}')
            if account in list(statistics_map.keys()):
                # print(f'合并 {account},{statistics_map[account][7].value}+{login_num},{statistics_map[account][8].value}+{replace_num}')
                statistics_map[account][7].value = statistics_map[account][7].value + login_num
                statistics_map[account][8].value = statistics_map[account][8].value + replace_num
            else:
                # print(f'新增 {account},{login_num},{replace_num}')
                statistics_map[account] = row


print(statistics_map['liuqinsgs1'][7].value, statistics_map['liuqinsgs1'][8].value)

# 保存合并结果
wb = Workbook()
ws = wb.active
ws.title = '统计信息'
for index, row in enumerate(statistics_map.values()):
    if row[0].value == 'liuqinsgs1':
        print(f'{row[7].value}, {row[8].value}')
    for col_index in range(0, 9 + 1):
        ws.cell(row=index + 1, column=col_index + 1, value=row[col_index].value)

print(f'统计信息总数: {len(statistics_map.values())}')
wb.save(f'{result_dir}/合并统计信息.xlsx')
