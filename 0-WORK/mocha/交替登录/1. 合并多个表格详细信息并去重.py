from openpyxl import load_workbook, Workbook

# 数据源路径
source_dir = 'C:/Users/hp/Desktop/交替登陆/需要删除新vpn'
# 处理结果存放路径
result_dir = source_dir + '/result'

result_detail_row_strings = []

for i in range(1, 24 + 1):
    excel_path = f'{source_dir}/{i}交替登录.xlsx'
    wb = load_workbook(excel_path)
    detail_sheet = wb['详细信息']
    detail_rows = list(detail_sheet.rows)

    for index, row in enumerate(detail_rows):
        if i == 1 or (i != 1 and index > 1):
            values = list(map(lambda c: c.value, row))
            if None not in values:
                join_str = '$'.join(values)
                print(join_str)
                result_detail_row_strings.append(join_str)

去重前大小 = len(result_detail_row_strings)
# 去重
result_detail_row_strings = list(set(result_detail_row_strings))
去重后大小 = len(result_detail_row_strings)

result_detail_rows = list(map(lambda rs: rs.split('$'), result_detail_row_strings))

result_wb = Workbook()
ws = result_wb.active
ws.title = '详细信息'
for index, row in enumerate(result_detail_rows):
    for col_index in range(0, 4 + 1):
        ws.cell(row=index + 1, column=col_index + 1, value=row[col_index])

result_wb.save(f'{result_dir}/合并详细信息.xlsx')

print(f'详细信息：去重前大小={去重前大小}, 去重后大小={去重后大小}')
