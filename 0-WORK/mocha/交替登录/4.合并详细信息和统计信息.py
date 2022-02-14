from openpyxl import load_workbook, Workbook

statistics_path = 'C:/Users/hp/Desktop/交替登陆/需要删除新vpn/result/合并统计信息.xlsx'
detail_path = 'C:/Users/hp/Desktop/交替登陆/需要删除新vpn/result/合并详细信息.xlsx'

statistics_book = load_workbook(statistics_path)
detail_book = load_workbook(detail_path)

statistics_sheet = statistics_book.active
detail_sheet = detail_book.active

statistics_rows = list(statistics_sheet.rows)
detail_rows = list(detail_sheet.rows)

wb = Workbook()
ws = wb.active
ws.title = '统计信息'
for index, row in enumerate(statistics_rows):
    for col_index in range(0, 9 + 1):
        ws.cell(row=index + 1, column=col_index + 1, value=row[col_index].value)

ws = wb.create_sheet('详细信息')
for index, row in enumerate(detail_rows):
    for col_index in range(0, 4 + 1):
        ws.cell(row=index + 1, column=col_index + 1, value=row[col_index].value)


wb.save('C:/Users/hp/Desktop/交替登陆/需要删除新vpn/result/合并24张表-详细信息已去重.xlsx')
