from selenium import webdriver
import time
import openpyxl
import pickle

browser = webdriver.Chrome()
browser.get("http://xxhapp.js.cmcc:30002/tokr-web/getuser?userid=822cdeeac1f3dd5294dc0c289a4ea0df")
# 浏览器最大化
#browser.maximize_window()
time.sleep(2)
# 填入用户名和密码
browser.find_element_by_name("userName").send_keys("okradmin")
time.sleep(1)
browser.find_element_by_name("password").send_keys("OKR@2018_admin")
time.sleep(1)
# 点击登录按钮
browser.find_element_by_name("submit").click()
time.sleep(6)
# 点击清空筛选
#browser.find_element_by_id("cleanCurFilter").click()
#time.sleep(6)

print("开始写入解析数据并写入到Excel文件")
# 打开 Excel 模板 C:\Users\hp\Desktop\OKR_list_export_template.xls
wb = openpyxl.load_workbook(r'C:\Users\hp\Desktop\OKR_list_export_template.xlsx')
ws = wb.worksheets[0]

workid_list = None
with open("work_id.pickle", "rb") as f:
    workid_list = pickle.load(f)

count = 2
# workid 查找完毕，遍历 workid，分析任务单详情页面
for workid in workid_list:
    # 打开任务单页面
    url = "http://xxhapp.js.cmcc:30002/tokr-web/fwd/corework/detail?workId=" + workid
    browser.get(url)
    # time.sleep(3)
    # browser.find_element_by_css_selector("a.cmcc_filters_toggle.moreFieldDown").click();time.sleep(1)

    while True:
        try:
            browser.execute_script("""
                d = document.getElementById("moreFieldCon");
                d.style = "display:block;";
            """)
            break
        except:
            pass

    task_detail_div = browser.find_element_by_id("moreFieldCon")
    tds = task_detail_div.find_elements_by_tag_name("td")

    print(f'{count-1} / {len(workid_list)} -> {workid} -> {browser.find_element_by_css_selector("h2.cmcc_title span").text}')

    # 将信息写入 Excel
    # workid
    ws['A' + str(count)].value = workid
    # 任务名称
    ws['B' + str(count)].value = browser.find_element_by_css_selector("h2.cmcc_title span").text
    # 任务状态
    ws['C' + str(count)].value = browser.find_element_by_css_selector("h2.cmcc_title i").text
    # 主办人
    ws['D' + str(count)].value = tds[1].text
    # 要求办结日期
    ws['E' + str(count)].value = tds[3].text
    # 交办人
    ws['F' + str(count)].value = tds[5].text
    # 汇报周期
    ws['G' + str(count)].value = tds[7].text
    # 主办部门
    ws['H' + str(count)].value = tds[9].text
    # 重要度
    ws['I' + str(count)].value = tds[11].text
    # 业务类型
    ws['J' + str(count)].value = tds[13].text
    # 任务类型
    ws['K' + str(count)].value = tds[15].text
    # 任务标签
    ws['L' + str(count)].value = tds[17].text
    # 任务模板
    ws['M' + str(count)].value = tds[19].text
    # 进度知悉人
    ws['N' + str(count)].value = tds[21].text
    # 协办人
    ws['O' + str(count)].value = tds[23].text
    # 协办部门
    ws['P' + str(count)].value = tds[25].text
    # 评价
    ws['Q' + str(count)].value = tds[27].text
    # 创建人
    ws['R' + str(count)].value = tds[29].text
    # 创建时间
    ws['S' + str(count)].value = tds[31].text
    # 任务编号
    ws['T' + str(count)].value = tds[33].text
    # 任务描述
    ws['U' + str(count)].value = tds[36].text

    # # 信息收集
    # info = ""
    # files = tds[38].find_elements_by_tag_name("b")
    # for file in files:
    #     info += file.text
    #     info += download_url + file.get_attribute("id")
    #     info += "\n"
    # ws['U' + str(count)].value = info
    # # 附件信息
    # info = ""
    # files = tds[40].find_elements_by_tag_name("b")
    # for file in files:
    #     info += file.text
    #     info += download_url + file.get_attribute("id")
    #     info += "\n"
    # ws['V' + str(count)].value = tds[40].text

    # print("导出进度：" + str((count / len(workid_list)) * 100) + "%")

    # browser.close()# 3秒后关闭
    count += 1

wb.save(r'C:\Users\hp\Desktop\result.xlsx')

print("导出完成！")
browser.execute_script("alert('导出完成！！！')")
time.sleep(30000)
browser.close()
