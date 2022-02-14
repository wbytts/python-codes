from selenium import webdriver
import time
import openpyxl
import pickle

workid_list = []
download_url = "http://xxhapp.js.cmcc:30002/tokr-web/fileresource/download?resourceId="
okr_home_url = 'http://xxhapp.js.cmcc:30002/tokr-web/'

browser = webdriver.Chrome()
browser.get("http://xxhapp.js.cmcc:30002/tokr-web/getuser?userid=822cdeeac1f3dd5294dc0c289a4ea0df")

# 浏览器最大化
browser.maximize_window()
time.sleep(2)
# 填入用户名和密码
browser.find_element_by_name("userName").send_keys("okradmin")
time.sleep(1)
browser.find_element_by_name("password").send_keys("OKR@2018_admin")
time.sleep(1)
# 点击登录按钮
browser.find_element_by_name("submit").click()
time.sleep(6)
# 点击清空筛选
browser.find_element_by_id("cleanCurFilter").click()
time.sleep(3)

# 读取要查找的名单
wb = openpyxl.load_workbook(r'F:\work\工作文档\OKR\21-03-22\2021-03-22 (1).xlsx')
ws = wb.active
task_num_list = []  # 任务单编号列表
for i in range(2, ws.max_row + 1):
    cell_task_num = ws.cell(row=i, column=1)
    cell_main = ws.cell(row=i, column=3)
    cell_title = ws.cell(row=i, column=2)
    task_num_list.append((cell_task_num.value, cell_main.value, cell_title.value))

task_map = {}
row_num = 2
# 遍历任务单编号，查看内容
for (task_num, zhuban, task_title) in task_num_list:
    # 输入任务单编号，搜索内容
    browser.find_element_by_css_selector("[name='search.key']").send_keys(task_num)
    time.sleep(2)
    browser.find_element_by_id('filterSearch').click()
    time.sleep(2)
    # 获取到搜索的任务结果
    task_openbutton_list = browser.find_elements_by_link_text("打开")
    if not task_openbutton_list: continue
    workids = [link.get_attribute("workid") for link in task_openbutton_list]
    for workid in workids:
        # 拼接任务单页面
        task_url = "http://xxhapp.js.cmcc:30002/tokr-web/fwd/corework/detail?workId=" + workid
        browser.get(task_url)
        time.sleep(3)
        browser.execute_script("""
                d = document.getElementById("moreFieldCon");
                d.style = "display:block;";
            """)
        task_detail_div = browser.find_element_by_id("moreFieldCon")
        tds = task_detail_div.find_elements_by_tag_name("td")
        if task_num == tds[33].text and zhuban == tds[1].text and browser.find_element_by_css_selector("h2.cmcc_title span").text == task_title:
            # 解析分配给的人员
            res = browser.find_elements_by_css_selector('#workTasks span.text-right')
            print(task_num + '\t:\t', end='')
            child_list = []
            for r in res:
                if '@' in r.text:
                    print(r.text.split('@')[0], '| ', end='')
                    child_list.append(r.text.split('@')[0])
            if task_num not in task_map:
                task_map[task_num] = {}
                task_map[task_num]['child_list'] = child_list[:]
                ws.cell(row=row_num, column=14, value=','.join(child_list))
                row_num += 1
            else:
                print(task_num, '#', '待确认')
            child_list.clear()
            print()
        else:
            pass

    # 回到主页，继续搜索
    browser.get(okr_home_url)
    time.sleep(3)
    # 点击清空筛选
    browser.find_element_by_id("cleanCurFilter").click()
    time.sleep(3)

with open('task_map.pickle', 'wb') as f:
    pickle.dump(task_map, f)

browser.close()  # 关闭浏览器
wb.save(r'F:\work\工作文档\OKR\21-03-22\2021-03-22-result.xlsx')
