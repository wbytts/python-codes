from selenium import webdriver
import time
import openpyxl
import pickle

browser = webdriver.Chrome()
browser.get("http://xxhapp.js.cmcc:30002/tokr-web/getuser?userid=822cdeeac1f3dd5294dc0c289a4ea0df")
# 浏览器最大化
# browser.maximize_window()
time.sleep(2)
# 填入用户名和密码
browser.find_element_by_name("userName").send_keys("okradmin")
time.sleep(1)
browser.find_element_by_name("password").send_keys("OKR@2018_admin")
time.sleep(1)
# 点击登录按钮
browser.find_element_by_name("submit").click()
time.sleep(6)
# 点击清空筛选
# browser.find_element_by_id("cleanCurFilter").click()
# time.sleep(6)

print("开始写入解析数据并写入到Excel文件")
# 打开 Excel 模板 C:\Users\hp\Desktop\OKR_list_export_template.xls
wb = openpyxl.load_workbook(r'C:\Users\hp\Desktop\OKR_list_export_template.xlsx')
ws = wb.worksheets[0]

workid_list = [
    'cw-93d563aa-9f5d-4063-b4a1-18d84217c6b2',
    'cw-9a592c55-2716-48dd-ba2d-f891c02e7255',
    'cw-8a1bd363-50f3-4afc-917c-cabfd9fe44c7',
    'cw-2725d372-884b-4ceb-8226-cb739bded1f9',
    'cw-06feb0e8-58c8-4722-a51d-5a868a1e8d24',
    'cw-3fc49ec1-54a1-4afe-9ea6-f004bbbff22c',
    'cw-0ac4a0b4-b3b6-4398-af0c-0bdf457fb98e',
    'cw-d7c44757-8649-46c6-a71e-3f60236d3d96',
    'cw-429be1e3-4af9-45b5-9aeb-57877f057ff6',
    'cw-666afc83-812a-4bc9-bf1f-05fd678ae72c',
    'cw-751fd6c5-5f8d-4336-b17d-d14faa18d9e5',
    'cw-03a0c736-adf8-487c-ac4c-4960fb04ef2b',
    'cw-16331120-bd2f-4470-8da3-3d609eb633db',
    'cw-c9a8285a-10ab-4984-97bf-a5f345cb316a',
    'cw-1394bbac-05dc-44b8-9e38-7ec9d9c76d25',
    'cw-c1501d2d-7727-4f5c-a3e9-ece098743182',
    'cw-fef7c4ab-04aa-40b6-bad3-8c4d71913eea',
    'cw-8f12dbed-2b0c-4515-8ea0-cf02a31e96c0',
    'cw-9b05e7b3-8daa-461e-bcdc-1e04598213e9',
    'cw-1b3ba668-4b14-4b3d-87e6-be7133be87c1',
    'cw-54ccb776-7369-49ee-96d0-1159466554da',
    'cw-072853aa-f732-4a17-b43e-8a5bf1bb3260',
    'cw-ed896a28-2c4c-4586-be25-fd48689af5ce',
    'cw-e51a8ec2-aba8-46f6-bed8-6b9528fcd2aa',
    'cw-9d8ba241-fe4e-4702-953b-e6b28b366e39',
    'cw-749b7a4f-e6d5-455c-ba96-835134985aa2',
    'cw-73a93b0e-4c40-43fd-8fed-4adede3e03dc',
    'cw-d20c1cae-a355-46e2-93ff-c65429380bd7',
    'cw-6e897b14-49e3-4ea6-86f4-5c36323f017b',
    'cw-b3f2be1b-481b-4f46-9326-9f71cc702da2',
    'cw-11aa22f0-fd83-4bef-91cb-56ff5276c88d',
    'cw-cd678636-6b38-4e36-a074-01f9d2a0f8b5',
    'cw-8bdc3f67-df9f-481f-9bc2-5251ebd1f1ca',
    'cw-c00b8d46-7dcf-4a22-8a0b-8e2902f0fe7f',
    'cw-b1b4224e-ff38-4627-ad08-c577b838b41f',
    'cw-28a7659e-e25a-4290-a3e7-44ee4d49d3f1',
    'cw-68fac4ac-1afe-41bb-b2ca-f940b4b52aec',
    'cw-2a0daa4e-ce01-475b-b8ee-a5dd5b920fb0',
    'cw-ef01a420-946d-498e-8264-a8b073c49fa4',
    'cw-85ffb2c6-d24e-4866-9f64-d8a21529280b',
    'cw-bdfc3c01-23c2-4946-9e31-5d712883ea12',
    'cw-2f9c5751-7030-4926-8e8f-e8d478ef31d7',
    'cw-ee27b3a9-ebcc-495d-a014-4eca2bab1fe5',
    'cw-b17e7006-040c-4454-97da-88e1f70f342c',
    'cw-58140005-675b-486b-808d-0a9561fe1353',
    'cw-380af94b-5fe7-4142-b473-b426f8e7fc7f',
    'cw-13f8fab2-d95b-474a-8d96-146fc45341e0',
    'cw-23d75fb7-18a1-4dbb-be19-711e14d5428a',
    'cw-8fee9006-77cc-4125-a424-ace53ad960ad',
    'cw-d563f0d1-4efc-4157-a625-901d99644138',
    'cw-89efe865-4fcb-4fc7-a0f7-ce93a7fbe834',
    'cw-fc2caa24-2c7b-4347-9caf-e64b97fed568',
    'cw-e114bfb4-4cb9-42da-af9b-1399eed51acd',
    'cw-83207101-cdf9-416b-a0bd-2b4378549e44',
    'cw-c499b546-2a63-43c8-872c-8b8d51ab25e9',
    'cw-195bf14b-f498-42af-a1e4-095640ff4ad9',
    'cw-ab846001-e655-4c3e-999d-cc76e5185da1',
    'cw-65150d40-4c58-4f5e-b4ef-371fbbda8520',
    'cw-d39d4636-110d-4c88-8fc6-97b2e1447468',
    'cw-c846013e-245e-448b-89e9-ee2bfed7cd36',
    'cw-8b625661-e6a6-4b95-85ec-cb1b1867abc5',
    'cw-52a6656b-2aa7-4bbb-a1aa-62fc17813c17',
    'cw-89d25a09-1889-42a4-a584-4bbcf51274f8',
    'cw-d6dcce21-3699-427f-b13e-76c19abfa3e2',
    'cw-5f976110-bf05-4829-ab75-605b79fc40fc',
    'cw-cb2f71aa-5620-4f84-8a5c-8e75788f7477',
    'cw-743b390f-e689-42cc-9c44-61229222db83',
    'cw-709d4bf4-2903-4aea-8e8e-bc1acba4b265',
    'cw-08085ed4-be38-432a-82ca-ab27fb7b3d38',
    'cw-16cc540b-53b3-42fe-8e34-1628344df5e2',
    'cw-3e412dc1-ea35-419b-8c7b-c6c564f6e5f4',
    'cw-5b01b31d-076f-4647-8580-99eb031fde81',
    'cw-2ff3c928-f46f-4148-b7d5-788a9911cab3',
    'cw-8b6d909d-2e97-47c4-90b5-36bd5c8bc021',
    'cw-15a1a0d1-4cc1-4ffe-ae94-afed717072c7',
    'cw-38f9a83d-802b-467e-8fc6-c9a0de8f9a9b',
    'cw-41dd71b3-eec9-43fd-bd69-ea918dfc5540',
    'cw-5e8502b4-f10b-4223-9ea7-daac52ee895e',
    'cw-c91faa17-dbc9-4eb9-a939-57b11d8dc04f',
    'cw-ecd6fbfc-f2bf-4ff4-a418-4a11ff7b2b28',
    'cw-b3295bb8-dfed-455f-918a-8a9f4725e230',
    'cw-36478e01-2af9-46da-9e79-875dc6ac5c49',
    'cw-713431b0-c6b1-4781-8fba-c63ab683e265',
    'cw-e4a7c4af-35eb-4dcf-8e49-cc479fcc7e14',
    'cw-5545098e-dac5-40e0-82c9-5c1b0669eebf',
    'cw-3d3cbd6b-a414-4325-aa6a-871b58838eb8',
    'cw-b608fe77-347c-4aa0-92b6-139208196b2e',
    'cw-fdaaecaf-b10f-426d-bf3e-610669cc6991',
    'cw-46980854-1bf1-4c8d-a633-1591a28d48cb',
    'cw-e48e7e62-d488-4445-9e14-0a055b119a76',
    'cw-2f1185b0-1a10-4bb1-be40-eeb1f6e2a8be',
    'cw-e51675f3-cdd2-4f38-bdc2-9d7c120cea8e',
    'cw-80d6daf3-6f7b-4ca6-8863-1e75e828b4a3',
    'cw-5e2447d3-6474-45f9-b06c-c8cb7e4ce6c8',
    'cw-13811a34-6aab-47e3-97b2-1202ab02e88a',
    'cw-587d647c-e098-451e-ae37-fcb018938ba4',
    'cw-a59f718a-9a03-493f-b9cf-75ac6f367bca',
    'cw-dd88bddd-9414-4f63-9004-628528d39e47',
    'cw-d79eb4a7-4ff3-4f9a-9002-96c6bff7894b',
    'cw-ca0842d4-367a-463f-b3ba-baed9ba0ad53',
    'cw-9022cf52-7e2a-49bf-986f-af4bf6b65ca0',
    'cw-de712a5b-71e4-41f4-8d25-09d3c231c9c6',
    'cw-9d41e088-42d2-45f0-866a-f33faca11c4f',
    'cw-e73692df-0d82-4b01-b7bd-6fc6fb748f0a',
    'cw-467323d8-6eae-44de-a548-2b6285ab7dcc',
    'cw-72d75988-d415-4505-bb85-1f02122b9a5a',
    'cw-9bde53f2-ed08-4fc4-a50c-8d7e0799dcca',
    'cw-a7ba24fe-06f8-481a-ae94-d32bf75c37f8',
    'cw-1af10c14-fde0-4393-bfd1-6967cd29084f',
    'cw-8adddf7d-0744-45f2-9390-281bd3067b4b',
    'cw-1f910d62-75f8-448c-ae19-24df4d3321f4',
    'cw-613cf5ca-acc8-4188-b1cd-db7eb51bc4fa',
    'cw-161cd2fa-39df-4387-9f98-3c9b6eee948e',
    'cw-a6130963-b915-4a8d-88dd-3c3219e34897',
    'cw-94be2f8b-ac8d-4ec0-a046-b88f1395e659',
    'cw-6c666cb2-5ef6-47ab-91ee-0184d5975907',
    'cw-8d1fcc23-a373-4144-8957-2bd156180e4c',
    'cw-6069cdd6-58e0-4a5c-972c-ebd0888ff6a1',
    'cw-3df9e66c-3e02-4f99-846e-cc46c43551b5',
    'cw-75e7607a-547c-4beb-9fc6-5b3f4949bb5d',
    'cw-48ce7d6a-a8c1-482e-9cd0-64c7dcdc148f',
    'cw-12bcf8b6-fe4b-4028-9e3b-1f2f2c16d101',
    'cw-ea73985c-a69d-46e3-898f-83fd5863f414',
    'cw-30c6bd50-3115-4fed-90fc-30980dfae5d2',
    'cw-ef19b394-944d-48da-8894-fb97a064fe1f',
    'cw-c2f84786-9d97-4350-aada-02a2d6bdf474',
    'cw-40dde065-17c9-42ca-9be4-ead49b678447',
    'cw-0a822a63-1dc2-40b4-9d04-1859fbb09fe3',
    'cw-fba76a54-65db-474e-be2b-5884e6a73340',
    'cw-a5b86351-e777-415d-a37a-85331ef3bb4d',
    'cw-ea4f74d1-44bc-40de-96b2-1652a594d3e4',
    'cw-fc87bdbd-e503-4f4b-a6dd-5d70b505f842',
    'cw-7aa77fd9-1c85-439d-a775-13a0df93cd68',
    'cw-c47c8b32-701b-44b7-a3aa-9654f3aa8fea',
    'cw-7033efaf-5281-4c34-9b9f-20a7c268fa81',
    'cw-cc519773-301e-438e-90d3-89152b1267d4',
    'cw-4b085753-64a2-4e19-a6dd-432750b54382',
    'cw-55cbf92c-561e-470e-bf26-f3c1fa43f922',
    'cw-a190262a-c9ef-42d5-bbe7-3f6ab4f52ee6',
    'cw-c0025c00-cfeb-4445-8ea4-1523461e5a1f',
    'cw-51d731ed-6605-40d5-a8a9-7764b19e4950',
    'cw-75af02b4-1fce-4535-8312-3abe257c83b9',
    'cw-1bc673fc-35ec-4b11-ab73-d0c8809fec7d',
    'cw-a8dbe1cc-1bd5-49e6-a0ff-7b1d292045dd',
    'cw-1121f3e6-ce9b-4f22-bc07-5cb18af72b95',
    'cw-3473deb6-1e03-4b6d-af42-9f7ac691eeca',
    'cw-a703578f-9937-444c-989f-fd4d97221356',
    'cw-d21fc7c7-d2c9-4a23-bcd2-43c7bb1485ef',
    'cw-04127ad5-8fd0-49bf-beba-b3dc3670e935',
    'cw-c3d98dbf-3bd8-4f62-ab4d-23c77be06d44',
    'cw-6e1ec908-6fb7-4134-97b8-504cdb05118b',
    'cw-b773c406-8cad-4652-ba2b-9f87e639714c',
    'cw-12be2793-90a7-48ca-a5c7-678c552f80e3',
    'cw-43d27bc0-34ea-47af-b9d4-23633f9d4983',
    'cw-956c2944-33eb-4226-a9d7-55a17a02aa9a',
    'cw-9224c7b0-a8a7-42df-ba27-451776fe3d30',
    'cw-5d0da36a-b8b6-4a3a-b43e-dba452429277',
    'cw-80f66703-34dc-49fb-885c-d4c8cafb2a01',
    'cw-f2c2e982-2d1e-4de1-a1ee-7f88bf55e838',
    'cw-7deebaad-0b28-49a4-8d58-0d6639571f29',
    'cw-bc82600c-4607-4f77-9c95-ed902cc1dcfd',
    'cw-f8b81a3a-1a08-43f9-ac08-6d2dcff9ff85',
]

count = 2
# workid 查找完毕，遍历 workid，分析任务单详情页面
for workid in workid_list:
    # 打开任务单页面
    url = "http://xxhapp.js.cmcc:30002/tokr-web/fwd/corework/detail?workId=" + workid
    browser.get(url)
    # time.sleep(3)
    # browser.find_element_by_css_selector("a.cmcc_filters_toggle.moreFieldDown").click();time.sleep(1)

    while True:
        try:
            browser.execute_script("""
                d = document.getElementById("moreFieldCon");
                d.style = "display:block;";
            """)
            break
        except:
            pass

    task_detail_div = browser.find_element_by_id("moreFieldCon")
    tds = task_detail_div.find_elements_by_tag_name("td")

    print(
        f'{count - 1} / {len(workid_list)} -> {workid} -> {browser.find_element_by_css_selector("h2.cmcc_title span").text}')

    # 将信息写入 Excel
    # workid
    ws['A' + str(count)].value = workid
    # 任务名称
    ws['B' + str(count)].value = browser.find_element_by_css_selector("h2.cmcc_title span").text
    # 任务状态
    ws['C' + str(count)].value = browser.find_element_by_css_selector("h2.cmcc_title i").text
    # 主办人
    ws['D' + str(count)].value = tds[1+4].text
    # 要求办结日期
    ws['E' + str(count)].value = tds[3+4].text
    # 交办人
    ws['F' + str(count)].value = tds[5+4].text
    # 汇报周期
    ws['G' + str(count)].value = tds[7+4].text
    # 主办部门
    ws['H' + str(count)].value = tds[9+4].text
    # 重要度
    ws['I' + str(count)].value = tds[11+4].text
    # 业务类型
    ws['J' + str(count)].value = tds[13+4].text
    # 任务类型
    ws['K' + str(count)].value = tds[15+4].text
    # 任务标签
    ws['L' + str(count)].value = tds[17+4].text
    # 任务模板
    ws['M' + str(count)].value = tds[19+4].text
    # 进度知悉人
    ws['N' + str(count)].value = tds[21+4].text
    # 协办人
    ws['O' + str(count)].value = tds[23+4].text
    # 协办部门
    ws['P' + str(count)].value = tds[25+4].text
    # 评价
    ws['Q' + str(count)].value = tds[27+4].text
    # 创建人
    ws['R' + str(count)].value = tds[29+4].text
    # 创建时间
    ws['S' + str(count)].value = tds[31+4].text
    # 任务编号
    ws['T' + str(count)].value = tds[33+4].text
    # 任务描述
    ws['U' + str(count)].value = tds[36+4].text

    # # 信息收集
    # info = ""
    # files = tds[38].find_elements_by_tag_name("b")
    # for file in files:
    #     info += file.text
    #     info += download_url + file.get_attribute("id")
    #     info += "\n"
    # ws['U' + str(count)].value = info
    # # 附件信息
    # info = ""
    # files = tds[40].find_elements_by_tag_name("b")
    # for file in files:
    #     info += file.text
    #     info += download_url + file.get_attribute("id")
    #     info += "\n"
    # ws['V' + str(count)].value = tds[40].text

    # print("导出进度：" + str((count / len(workid_list)) * 100) + "%")

    # browser.close()# 3秒后关闭
    count += 1

wb.save(r'C:\Users\hp\Desktop\result.xlsx')

print("导出完成！")
browser.execute_script("alert('导出完成！！！')")
time.sleep(30000)
browser.close()
