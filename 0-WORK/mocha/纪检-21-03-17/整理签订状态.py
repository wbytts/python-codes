import openpyxl as opx

# 原始待发送名单
send_list = opx.load_workbook(r'F:\work\工作文档\纪检\纪检-21-03-17\公司员工杜绝酒驾醉驾承诺书-方宁的模板.xlsx')
send_sheet = send_list.active
# 添加的名单
# 已签订的名单
finish_book = opx.load_workbook(r'F:\work\工作文档\纪检\纪检-21-03-17\yq0402.xlsx')
finish_sheet = finish_book.active

need_send_list = []
# 收集名单中需要发送的人
for i in range(2, send_sheet.max_row + 1):
    cell = send_sheet.cell(row=i, column=2)
    if cell.value == '查询不到OA账号':
        send_sheet.cell(row=i, column=4, value='查询不到OA账号')
        continue
    elif cell.value is None:
        send_sheet.cell(row=i, column=4, value='离职')
        continue
    else:
        need_send_list.append(cell.value.split('@')[0])

# need_send_list.append('jwangxiaoxing')
# need_send_list.append('wangyixuan')


# 统计已签订的账号
yiq_list = []
for i in range(2, finish_sheet.max_row + 1):
    cell = finish_sheet.cell(row=i, column=1)
    yiq_list.append(cell.value.lower())

print(yiq_list)

# 遍历名单，标记已签订人员  “已签”
for i in range(2, send_sheet.max_row + 1):
    cell = send_sheet.cell(row=i, column=2)
    if send_sheet.cell(row=i, column=4).value == '未签':
        oa = None
        if '@' not in cell.value:
            oa = cell.value
        else:
            oa = send_sheet.cell(row=i, column=2).value.split('@')[0].lower()

        # if ('j' + oa) in yiq_list:
        #     send_sheet.cell(row=i, column=4, value='已签')
        #     continue

        if oa in yiq_list:
            send_sheet.cell(row=i, column=4, value='已签')
        # else:
        #     if oa in ['wangxylyg', 'gutongyun', 'zhanglei2', 'yinxinyc', 'lishanshan']:
        #         send_sheet.cell(row=i, column=4, value='未签')
        #     else:
        #         send_sheet.cell(row=i, column=4, value='未签')

send_list.save(r'F:\work\工作文档\纪检\纪检-21-03-17\result.xlsx')
