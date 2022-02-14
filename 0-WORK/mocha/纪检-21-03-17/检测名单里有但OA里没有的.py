from openpyxl import load_workbook

wb1 = load_workbook(r'F:\data\excel\20210316全省人员名单-已反馈.xlsx')
wb2 = load_workbook(r'F:\data\excel\OA-v3ry.xlsx')
ws1 = wb1.active
ws2 = wb2.active

all_accounts = []
oa_accounts = []

for i in range(1, ws2.max_row+1):
    cell = ws2.cell(row=i, column=2)
    oa_accounts.append(cell.value.lower())

for i in range(2, ws1.max_row+1):
    cell = ws1.cell(row=i, column=2)
    oa = cell.value.split('@')[0].lower()
    if oa.lower() not in oa_accounts:
        ws1.cell(row=i, column=4).value = '#'

wb1.save(r'C:\Users\hp\Desktop\纪检-21-03-17\名单-not-in-oa.xlsx')
