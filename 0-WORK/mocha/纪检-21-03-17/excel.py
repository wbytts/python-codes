from openpyxl import load_workbook

wb1 = load_workbook(r'F:\data\excel\20210316全省人员名单-已反馈.xlsx')
wb2 = load_workbook(r'F:\data\excel\OA-v3ry.xlsx')
ws1 = wb1.active
ws2 = wb2.active

all_accounts = []
oa_accounts = []

# 获取所有 value（生成器）
for row in ws1.values:
    oa = row[1].split('@')[0].lower()
    all_accounts.append(oa)
all_accounts.pop(0)

for i in range(2, ws2.max_row+1):
    oa = ws2.cell(row=i, column=2).value
    # 忽略大小写检测 OA中的账号，在全省名单里有没有
    if oa.lower() not in all_accounts:
        ws2.cell(row=i, column=4, value='#')

wb2.save(r'C:\Users\hp\Desktop\纪检-21-03-17\OA-v3ry-mark.xlsx')
