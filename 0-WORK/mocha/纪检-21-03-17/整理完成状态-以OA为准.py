import openpyxl as opx

# OA人员表
oa_book = opx.load_workbook(r'F:\work\工作文档\纪检\纪检-21-03-17\OA-v3ry-删除.xlsx')
oa_sheet = oa_book.active
# 数据库导出的已签名单
finish_book = opx.load_workbook(r'F:\work\工作文档\纪检\纪检-21-03-17\1604yiq.xlsx')
finish_sheet = finish_book.active

oa_list = []  # oa账户列表
finish_list = []  # 已签列表

# 从数据库导出的表格中，统计已签人员
for i in range(2, finish_sheet.max_row + 1):
    cell = finish_sheet.cell(row=i, column=1)
    finish_list.append(cell.value)

finish_count = 0  # 统计已签人数
# 遍历OA表，标记签订状态
for i in range(2, oa_sheet.max_row + 1):
    cell = oa_sheet.cell(row=i, column=2)
    if cell.value in finish_list:
        finish_count += 1
        oa_sheet.cell(row=i, column=4, value='已签')

print('已签人数: ', finish_count)
# 保存结果表格
oa_book.save(r'F:\work\工作文档\纪检\result.xlsx')
